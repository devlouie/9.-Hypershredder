#!/usr/bin/env python3

import os
import re
import sys
import datetime
from pathlib import Path
import PyPDF2
from docx import Document
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from xml.dom import minidom
import logging
from typing import Optional, Dict, Any

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DocumentProcessor:
    def __init__(self, input_dir: str, output_file: str):
        self.input_dir = Path(input_dir)
        self.output_file = Path(output_file)
        self.root = ET.Element("corpus")
        self.root.set("timestamp", datetime.datetime.now().isoformat())
        
    def clean_text(self, text: str) -> str:
        """Clean and normalize text content."""
        if not text:
            return ""
        # Normalize whitespace
        text = re.sub(r'\s+', ' ', text)
        # Remove control characters
        text = ''.join(char for char in text if ord(char) >= 32 or char in '\n\t')
        return text.strip()

    def process_pdf(self, file_path: Path) -> Optional[str]:
        """Extract text from PDF files."""
        try:
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text = []
                for page_num in range(len(reader.pages)):
                    text.append(reader.pages[page_num].extract_text())
                return '\n'.join(text)
        except Exception as e:
            logger.error(f"Error processing PDF {file_path}: {str(e)}")
            return None

    def process_docx(self, file_path: Path) -> Optional[str]:
        """Extract text from DOCX files."""
        try:
            doc = Document(file_path)
            text = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text.append(paragraph.text)
            return '\n'.join(text)
        except Exception as e:
            logger.error(f"Error processing DOCX {file_path}: {str(e)}")
            return None

    def process_xlsx(self, file_path: Path) -> Optional[str]:
        """Extract text from XLSX files."""
        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            text = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text.append(f"<worksheet name='{sheet}'>")
                for row in ws.iter_rows():
                    row_text = [str(cell.value) if cell.value is not None else '' for cell in row]
                    if any(row_text):  # Only include non-empty rows
                        text.append('\t'.join(row_text))
                text.append("</worksheet>")
            return '\n'.join(text)
        except Exception as e:
            logger.error(f"Error processing XLSX {file_path}: {str(e)}")
            return None

    def create_document_element(self, file_path: Path, content: str) -> ET.Element:
        """Create an XML element for a document."""
        doc = ET.SubElement(self.root, "document")
        
        # Add metadata
        metadata = ET.SubElement(doc, "metadata")
        ET.SubElement(metadata, "filename").text = file_path.name
        ET.SubElement(metadata, "type").text = file_path.suffix[1:].lower()
        ET.SubElement(metadata, "path").text = str(file_path.relative_to(self.input_dir))
        ET.SubElement(metadata, "timestamp").text = datetime.datetime.fromtimestamp(
            file_path.stat().st_mtime).isoformat()
        
        # Add content
        content_elem = ET.SubElement(doc, "content")
        content_elem.text = self.clean_text(content)
        
        return doc

    def process_file(self, file_path: Path) -> None:
        """Process a single file based on its extension."""
        logger.info(f"Processing {file_path}")
        
        content = None
        ext = file_path.suffix.lower()
        
        if ext == '.pdf':
            content = self.process_pdf(file_path)
        elif ext == '.docx':
            content = self.process_docx(file_path)
        elif ext == '.xlsx':
            content = self.process_xlsx(file_path)
        elif ext == '.txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        
        if content:
            self.create_document_element(file_path, content)

    def process_directory(self) -> None:
        """Process all supported files in the directory."""
        supported_extensions = {'.pdf', '.docx', '.xlsx', '.txt'}
        
        for file_path in self.input_dir.rglob('*'):
            if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                self.process_file(file_path)

    def save_output(self) -> None:
        """Save the processed documents to an XML file."""
        # Create a pretty-printed XML string
        xml_str = minidom.parseString(ET.tostring(self.root)).toprettyxml(indent="  ")
        
        # Save to file
        with open(self.output_file, 'w', encoding='utf-8') as f:
            f.write(xml_str)
        
        logger.info(f"Output saved to {self.output_file}")

def main():
    if len(sys.argv) != 3:
        print("Usage: python document_extractor.py <input_directory> <output_file>")
        sys.exit(1)
    
    input_dir = sys.argv[1]
    output_file = sys.argv[2]
    
    processor = DocumentProcessor(input_dir, output_file)
    processor.process_directory()
    processor.save_output()

if __name__ == "__main__":
    main() 