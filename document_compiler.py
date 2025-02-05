#!/usr/bin/env python3

import os
import re
import sys
import datetime
from pathlib import Path
from typing import Optional, Dict, Any, Tuple, List
import logging
from PIL import Image
import io
import PyPDF2
from docx import Document
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import pandas as pd
from document_utils import (
    optimize_image,
    extract_table_from_docx,
    create_reportlab_table,
    extract_images_from_pdf,
    extract_tables_from_pdf,
    clean_table_data
)

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DocumentCompiler:
    def __init__(self, input_dir: str, output_file: str):
        self.input_dir = Path(input_dir)
        self.output_file = Path(output_file)
        self.styles = getSampleStyleSheet()
        self.story = []  # Will hold all elements for the PDF
        
        # Modify existing styles
        self.styles['Heading1'].fontSize = 16
        self.styles['Heading1'].spaceAfter = 20
        
        # Create custom styles
        self.styles.add(ParagraphStyle(
            name='CustomMetadata',
            parent=self.styles['Normal'],
            fontSize=8,
            textColor=colors.gray,
            spaceAfter=12
        ))
        
        self.styles.add(ParagraphStyle(
            name='TableCaption',
            parent=self.styles['Normal'],
            fontSize=9,
            textColor=colors.gray,
            alignment=1,  # Center
            spaceAfter=6
        ))

    def clean_text(self, text: str) -> str:
        """Clean and normalize text content."""
        if not text:
            return ""
            
        try:
            # Replace common Unicode characters
            replacements = {
                '\u2013': '-',    # en dash
                '\u2014': '-',    # em dash
                '\u2018': "'",    # left single quote
                '\u2019': "'",    # right single quote
                '\u201c': '"',    # left double quote
                '\u201d': '"',    # right double quote
                '\u00f3': 'o',    # ó
                '\u00a0': ' ',    # non-breaking space
                '\u2022': '*',    # bullet point
                '\u00ae': '(R)',  # registered trademark
                '\u2122': '(TM)', # trademark
                '\u00a9': '(c)',  # copyright
            }
            
            # Replace special characters
            for old, new in replacements.items():
                text = text.replace(old, new)
            
            # Remove other non-printable characters
            text = ''.join(char for char in text if ord(char) >= 32 or char in '\n\t')
            
            # Normalize whitespace
            text = re.sub(r'\s+', ' ', text)
            
            return text.strip()
        except Exception as e:
            logger.error(f"Error cleaning text: {str(e)}")
            # Return a safe version of the text
            return ''.join(char for char in text if ord(char) < 128)

    def create_paragraph(self, text: str, style: str = 'Normal') -> Optional[Paragraph]:
        """Create a paragraph with error handling."""
        try:
            cleaned_text = self.clean_text(text)
            # Split into smaller chunks if text is too long
            max_chunk = 2000  # Reduced chunk size for safety
            
            if len(cleaned_text) <= max_chunk:
                return Paragraph(cleaned_text, self.styles[style])
            
            # Split long text into paragraphs
            chunks = []
            words = cleaned_text.split()
            current_chunk = []
            current_length = 0
            
            for word in words:
                if current_length + len(word) + 1 > max_chunk:
                    chunks.append(' '.join(current_chunk))
                    current_chunk = [word]
                    current_length = len(word)
                else:
                    current_chunk.append(word)
                    current_length += len(word) + 1
            
            if current_chunk:
                chunks.append(' '.join(current_chunk))
            
            return [Paragraph(chunk, self.styles[style]) for chunk in chunks]
            
        except Exception as e:
            logger.error(f"Error creating paragraph: {str(e)}")
            return None

    def add_metadata_block(self, file_path: Path):
        """Add metadata information to the PDF."""
        metadata_text = [
            f"File: {file_path.name}",
            f"Type: {file_path.suffix[1:].upper()}",
            f"Path: {file_path.relative_to(self.input_dir)}",
            f"Last Modified: {datetime.datetime.fromtimestamp(file_path.stat().st_mtime).isoformat()}"
        ]
        
        for line in metadata_text:
            self.story.append(Paragraph(line, self.styles['CustomMetadata']))
        self.story.append(Spacer(1, 12))

    def process_pdf(self, file_path: Path) -> None:
        """Extract content from PDF files with enhanced image and table support."""
        try:
            # Add document title
            title_para = self.create_paragraph(f"Document: {file_path.name}", 'Heading1')
            if title_para:
                self.story.append(title_para)
            self.add_metadata_block(file_path)
            
            # Extract and process images
            images = extract_images_from_pdf(str(file_path))
            for idx, (image_bytes, image_type) in enumerate(images, 1):
                try:
                    optimized_image = optimize_image(image_bytes)
                    if optimized_image:
                        img = RLImage(io.BytesIO(optimized_image), width=6*inch, height=4*inch, kind='proportional')
                        self.story.append(img)
                        self.story.append(Paragraph(f"Image {idx} from PDF", self.styles['TableCaption']))
                        self.story.append(Spacer(1, 12))
                except Exception as img_e:
                    logger.error(f"Error processing image {idx} from PDF: {str(img_e)}")
            
            # Extract and process tables
            tables = extract_tables_from_pdf(str(file_path))
            for idx, df in enumerate(tables, 1):
                try:
                    cleaned_df = clean_table_data(df)
                    table_data = [cleaned_df.columns.tolist()] + cleaned_df.values.tolist()
                    table = create_reportlab_table(table_data)
                    if table:
                        self.story.append(table)
                        self.story.append(Paragraph(f"Table {idx} from PDF", self.styles['TableCaption']))
                        self.story.append(Spacer(1, 12))
                except Exception as table_e:
                    logger.error(f"Error processing table {idx} from PDF: {str(table_e)}")
            
            # Process text content
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page_num in range(len(reader.pages)):
                    try:
                        text = reader.pages[page_num].extract_text()
                        if text.strip():
                            page_header = f"Page {page_num + 1}"
                            header_para = self.create_paragraph(page_header, 'CustomMetadata')
                            if header_para:
                                self.story.append(header_para)
                            
                            content_para = self.create_paragraph(text)
                            if isinstance(content_para, list):
                                self.story.extend(content_para)
                            elif content_para:
                                self.story.append(content_para)
                            
                            self.story.append(Spacer(1, 12))
                    except Exception as page_e:
                        logger.error(f"Error processing page {page_num + 1} of PDF {file_path}: {str(page_e)}")
                        continue
        except Exception as e:
            logger.error(f"Error processing PDF {file_path}: {str(e)}")

    def process_docx(self, file_path: Path) -> None:
        """Extract content from DOCX files with enhanced table support."""
        try:
            # Add document title
            title_para = self.create_paragraph(f"Document: {file_path.name}", 'Heading1')
            if title_para:
                self.story.append(title_para)
            self.add_metadata_block(file_path)
            
            doc = Document(file_path)
            
            # Process tables first
            for idx, table in enumerate(doc.tables, 1):
                try:
                    table_data = extract_table_from_docx(table)
                    if table_data:
                        rl_table = create_reportlab_table(table_data)
                        if rl_table:
                            self.story.append(rl_table)
                            self.story.append(Paragraph(f"Table {idx} from document", self.styles['TableCaption']))
                            self.story.append(Spacer(1, 12))
                except Exception as table_e:
                    logger.error(f"Error processing table {idx} in DOCX {file_path}: {str(table_e)}")
            
            # Process paragraphs and images
            for element in doc.element.body:
                if element.tag.endswith('p'):
                    # Process paragraph
                    para = doc.element.xpath('//w:p[count(.)=1]')[0]
                    if para.text.strip():
                        try:
                            para_obj = self.create_paragraph(para.text)
                            if isinstance(para_obj, list):
                                self.story.extend(para_obj)
                            elif para_obj:
                                self.story.append(para_obj)
                        except Exception as para_e:
                            logger.error(f"Error processing paragraph in DOCX {file_path}: {str(para_e)}")
                
                elif element.tag.endswith('r'):
                    # Process run (might contain images)
                    for rel in doc.part.rels.values():
                        if "image" in rel.target_ref:
                            try:
                                image_bytes = rel.target_part.blob
                                optimized_image = optimize_image(image_bytes)
                                if optimized_image:
                                    img = RLImage(io.BytesIO(optimized_image), width=6*inch, height=4*inch, kind='proportional')
                                    self.story.append(img)
                                    self.story.append(Spacer(1, 12))
                            except Exception as img_e:
                                logger.error(f"Error processing image in DOCX {file_path}: {str(img_e)}")
            
        except Exception as e:
            logger.error(f"Error processing DOCX {file_path}: {str(e)}")

    def process_xlsx(self, file_path: Path) -> None:
        """Extract content from XLSX files with enhanced table support."""
        try:
            # Add document title
            self.story.append(Paragraph(f"Document: {file_path.name}", self.styles['Heading1']))
            self.add_metadata_block(file_path)
            
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self.story.append(Paragraph(f"Sheet: {sheet_name}", self.styles['Heading2']))
                
                # Convert worksheet to DataFrame for better processing
                data = []
                for row in ws.iter_rows():
                    row_data = [str(cell.value) if cell.value is not None else '' for cell in row]
                    if any(row_data):  # Only include non-empty rows
                        data.append(row_data)
                
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    cleaned_df = clean_table_data(df)
                    table_data = [cleaned_df.columns.tolist()] + cleaned_df.values.tolist()
                    table = create_reportlab_table(table_data)
                    if table:
                        self.story.append(table)
                        self.story.append(Spacer(1, 12))
                
        except Exception as e:
            logger.error(f"Error processing XLSX {file_path}: {str(e)}")

    def process_image(self, file_path: Path) -> None:
        """Process and include image files with optimization."""
        try:
            # Add document title
            self.story.append(Paragraph(f"Image: {file_path.name}", self.styles['Heading1']))
            self.add_metadata_block(file_path)
            
            with open(file_path, 'rb') as f:
                image_data = f.read()
            
            optimized_image = optimize_image(image_data)
            if optimized_image:
                img = RLImage(io.BytesIO(optimized_image), width=6*inch, height=4*inch, kind='proportional')
                self.story.append(img)
                self.story.append(Spacer(1, 12))
        except Exception as e:
            logger.error(f"Error processing image {file_path}: {str(e)}")

    def process_file(self, file_path: Path) -> None:
        """Process a single file based on its extension."""
        logger.info(f"Processing {file_path}")
        
        ext = file_path.suffix.lower()
        
        if ext == '.pdf':
            self.process_pdf(file_path)
        elif ext == '.docx':
            self.process_docx(file_path)
        elif ext == '.xlsx':
            self.process_xlsx(file_path)
        elif ext in {'.jpg', '.jpeg', '.png', '.gif'}:
            self.process_image(file_path)
        elif ext == '.txt':
            # Add document title
            self.story.append(Paragraph(f"Document: {file_path.name}", self.styles['Heading1']))
            self.add_metadata_block(file_path)
            
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                if content.strip():
                    para = self.create_paragraph(content)
                    if isinstance(para, list):
                        self.story.extend(para)
                    elif para:
                        self.story.append(para)
        
        # Add a page break between documents
        self.story.append(Spacer(1, 20))

    def process_directory(self) -> None:
        """Process all supported files in the directory."""
        supported_extensions = {'.pdf', '.docx', '.xlsx', '.txt', '.jpg', '.jpeg', '.png', '.gif'}
        
        # Add title page
        self.story.append(Paragraph("Document Compilation", self.styles['Title']))
        self.story.append(Paragraph(
            f"Generated on {datetime.datetime.now().strftime('%A, %B %d, %Y at %I:%M:%S %p')}",
            self.styles['Normal']
        ))
        self.story.append(Paragraph(
            f"Source Directory: {self.input_dir}",
            self.styles['Normal']
        ))
        self.story.append(Spacer(1, 30))
        
        # Process all files
        for file_path in sorted(self.input_dir.rglob('*')):
            if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                self.process_file(file_path)

    def save_output(self) -> None:
        """Generate the final PDF."""
        doc = SimpleDocTemplate(
            str(self.output_file),
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        try:
            doc.build(self.story)
            logger.info(f"Output saved to {self.output_file}")
        except Exception as e:
            logger.error(f"Error generating PDF: {str(e)}")

def main():
    if len(sys.argv) != 2:
        print("Usage: python document_compiler.py <input_directory>")
        sys.exit(1)
    
    input_dir = sys.argv[1]
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f"compiled_docs_{timestamp}.pdf"
    
    processor = DocumentCompiler(input_dir, output_file)
    processor.process_directory()
    processor.save_output()

if __name__ == "__main__":
    main() 