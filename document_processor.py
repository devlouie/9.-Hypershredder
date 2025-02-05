"""
Document Processor
================

This module handles the processing and compilation of various document types into
a single, well-formatted PDF file.

Key Features:
------------
1. Multi-format document support
2. PDF compilation
3. Image optimization
4. Table extraction and formatting
5. Metadata preservation
"""

import os
import re
import sys
import datetime
from pathlib import Path
from typing import Optional, Dict, Any, Tuple, List
import logging
from PIL import Image
import io
import PyPDF2
from docx import Document
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import pandas as pd
from document_utils import (
    optimize_image,
    extract_table_from_docx,
    create_reportlab_table,
    extract_images_from_pdf,
    extract_tables_from_pdf,
    clean_table_data
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DocumentProcessor:
    """
    A class to process and compile various document types into a single PDF.
    
    This class handles:
    1. Document type detection
    2. Content extraction
    3. PDF compilation
    4. Formatting and styling
    """
    
    def __init__(self, input_dir: str, output_file: str):
        """
        Initialize the document processor.
        
        Args:
            input_dir (str): Directory containing input documents
            output_file (str): Path for the output PDF file
        """
        self.input_dir = Path(input_dir)
        self.output_file = Path(output_file)
        self.story = []
        self._setup_styles()
        
    def _setup_styles(self):
        """Set up document styles for PDF generation."""
        styles = getSampleStyleSheet()
        self.styles = {
            'Title': styles['Title'],
            'Heading1': styles['Heading1'],
            'Normal': styles['Normal'],
            'Code': styles['Code'],
            'TableCaption': ParagraphStyle(
                'TableCaption',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey
            ),
            'CustomMetadata': ParagraphStyle(
                'CustomMetadata',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey
            )
        }
        
    def clean_text(self, text: str) -> str:
        """
        Clean and normalize text content.
        
        Args:
            text (str): Input text
            
        Returns:
            str: Cleaned text
        """
        if not text:
            return ""
        # Normalize whitespace
        text = re.sub(r'\s+', ' ', text)
        # Remove control characters
        text = ''.join(char for char in text if ord(char) >= 32 or char in '\n\t')
        return text.strip()

    def create_paragraph(self, text: str, style: str = 'Normal') -> Optional[Paragraph]:
        """
        Create a paragraph with error handling.
        
        Args:
            text (str): Paragraph text
            style (str): Style name
            
        Returns:
            Optional[Paragraph]: Created paragraph or None if error
        """
        try:
            cleaned_text = self.clean_text(text)
            # Split into smaller chunks if text is too long
            max_chunk = 2000  # Reduced chunk size for safety
            
            if len(cleaned_text) <= max_chunk:
                return Paragraph(cleaned_text, self.styles[style])
            
            # Split long text into paragraphs
            chunks = []
            words = cleaned_text.split()
            current_chunk = []
            current_length = 0
            
            for word in words:
                if current_length + len(word) + 1 > max_chunk:
                    chunks.append(' '.join(current_chunk))
                    current_chunk = [word]
                    current_length = len(word)
                else:
                    current_chunk.append(word)
                    current_length += len(word) + 1
            
            if current_chunk:
                chunks.append(' '.join(current_chunk))
            
            return [Paragraph(chunk, self.styles[style]) for chunk in chunks]
            
        except Exception as e:
            logger.error(f"Error creating paragraph: {str(e)}")
            return None

    def add_metadata_block(self, file_path: Path):
        """
        Add metadata information to the PDF.
        
        Args:
            file_path (Path): Path to the source file
        """
        metadata_text = [
            f"File: {file_path.name}",
            f"Type: {file_path.suffix[1:].upper()}",
            f"Path: {file_path.relative_to(self.input_dir)}",
            f"Last Modified: {datetime.datetime.fromtimestamp(file_path.stat().st_mtime).isoformat()}"
        ]
        
        for line in metadata_text:
            self.story.append(Paragraph(line, self.styles['CustomMetadata']))
        self.story.append(Spacer(1, 12))

    def process_pdf(self, file_path: Path) -> None:
        """
        Process a PDF file.
        
        Args:
            file_path (Path): Path to the PDF file
        """
        try:
            # Add document title
            title_para = self.create_paragraph(f"Document: {file_path.name}", 'Heading1')
            if title_para:
                self.story.append(title_para)
            self.add_metadata_block(file_path)
            
            # Extract and process images
            images = extract_images_from_pdf(str(file_path))
            for idx, (image_bytes, image_type) in enumerate(images, 1):
                try:
                    optimized_image = optimize_image(image_bytes)
                    if optimized_image:
                        img = RLImage(io.BytesIO(optimized_image), width=6*inch, height=4*inch, kind='proportional')
                        self.story.append(img)
                        self.story.append(Paragraph(f"Image {idx} from PDF", self.styles['TableCaption']))
                        self.story.append(Spacer(1, 12))
                except Exception as img_e:
                    logger.error(f"Error processing image {idx} from PDF: {str(img_e)}")
            
            # Extract and process tables
            tables = extract_tables_from_pdf(str(file_path))
            for idx, df in enumerate(tables, 1):
                try:
                    cleaned_df = clean_table_data(df)
                    table_data = [cleaned_df.columns.tolist()] + cleaned_df.values.tolist()
                    table = create_reportlab_table(table_data)
                    if table:
                        self.story.append(table)
                        self.story.append(Paragraph(f"Table {idx} from PDF", self.styles['TableCaption']))
                        self.story.append(Spacer(1, 12))
                except Exception as table_e:
                    logger.error(f"Error processing table {idx} from PDF: {str(table_e)}")
            
            # Process text content
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page_num in range(len(reader.pages)):
                    try:
                        text = reader.pages[page_num].extract_text()
                        if text.strip():
                            page_header = f"Page {page_num + 1}"
                            header_para = self.create_paragraph(page_header, 'CustomMetadata')
                            if header_para:
                                self.story.append(header_para)
                            
                            content_para = self.create_paragraph(text)
                            if isinstance(content_para, list):
                                self.story.extend(content_para)
                            elif content_para:
                                self.story.append(content_para)
                            
                            self.story.append(Spacer(1, 12))
                    except Exception as page_e:
                        logger.error(f"Error processing page {page_num + 1} of PDF {file_path}: {str(page_e)}")
                        continue
        except Exception as e:
            logger.error(f"Error processing PDF {file_path}: {str(e)}")

    def process_docx(self, file_path: Path) -> None:
        """
        Process a Word document.
        
        Args:
            file_path (Path): Path to the DOCX file
        """
        try:
            # Add document title
            title_para = self.create_paragraph(f"Document: {file_path.name}", 'Heading1')
            if title_para:
                self.story.append(title_para)
            self.add_metadata_block(file_path)
            
            doc = Document(file_path)
            
            # Extract and process tables
            for idx, table in enumerate(doc.tables, 1):
                try:
                    df = extract_table_from_docx(table)
                    if not df.empty:
                        cleaned_df = clean_table_data(df)
                        table_data = [cleaned_df.columns.tolist()] + cleaned_df.values.tolist()
                        reportlab_table = create_reportlab_table(table_data)
                        if reportlab_table:
                            self.story.append(reportlab_table)
                            self.story.append(Paragraph(f"Table {idx}", self.styles['TableCaption']))
                            self.story.append(Spacer(1, 12))
                except Exception as table_e:
                    logger.error(f"Error processing table {idx} in DOCX {file_path}: {str(table_e)}")
            
            # Process paragraphs and images
            for element in doc.element.body:
                if element.tag.endswith('p'):
                    # Process paragraph
                    para = doc.element.xpath('//w:p[count(.)=1]')[0]
                    if para.text.strip():
                        try:
                            para_obj = self.create_paragraph(para.text)
                            if isinstance(para_obj, list):
                                self.story.extend(para_obj)
                            elif para_obj:
                                self.story.append(para_obj)
                        except Exception as para_e:
                            logger.error(f"Error processing paragraph in DOCX {file_path}: {str(para_e)}")
                
                elif element.tag.endswith('r'):
                    # Process run (might contain images)
                    for rel in doc.part.rels.values():
                        if "image" in rel.target_ref:
                            try:
                                image_bytes = rel.target_part.blob
                                optimized_image = optimize_image(image_bytes)
                                if optimized_image:
                                    img = RLImage(io.BytesIO(optimized_image), width=6*inch, height=4*inch, kind='proportional')
                                    self.story.append(img)
                                    self.story.append(Spacer(1, 12))
                            except Exception as img_e:
                                logger.error(f"Error processing image in DOCX {file_path}: {str(img_e)}")
            
        except Exception as e:
            logger.error(f"Error processing DOCX {file_path}: {str(e)}")

    def process_xlsx(self, file_path: Path) -> None:
        """
        Process an Excel file.
        
        Args:
            file_path (Path): Path to the XLSX file
        """
        try:
            # Add document title
            title_para = self.create_paragraph(f"Document: {file_path.name}", 'Heading1')
            if title_para:
                self.story.append(title_para)
            self.add_metadata_block(file_path)
            
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Add sheet header
                self.story.append(Paragraph(f"Sheet: {sheet_name}", self.styles['Heading1']))
                
                # Convert sheet to DataFrame
                data = []
                for row in ws.iter_rows():
                    row_data = [str(cell.value) if cell.value is not None else '' for cell in row]
                    if any(row_data):  # Only include non-empty rows
                        data.append(row_data)
                
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    cleaned_df = clean_table_data(df)
                    table_data = [cleaned_df.columns.tolist()] + cleaned_df.values.tolist()
                    table = create_reportlab_table(table_data)
                    if table:
                        self.story.append(table)
                        self.story.append(Spacer(1, 12))
                
            wb.close()
            
        except Exception as e:
            logger.error(f"Error processing XLSX {file_path}: {str(e)}")

    def process_image(self, file_path: Path) -> None:
        """
        Process an image file.
        
        Args:
            file_path (Path): Path to the image file
        """
        try:
            # Add document title
            self.story.append(Paragraph(f"Image: {file_path.name}", self.styles['Heading1']))
            self.add_metadata_block(file_path)
            
            with open(file_path, 'rb') as f:
                image_data = f.read()
            
            optimized_image = optimize_image(image_data)
            if optimized_image:
                img = RLImage(io.BytesIO(optimized_image), width=6*inch, height=4*inch, kind='proportional')
                self.story.append(img)
                self.story.append(Spacer(1, 12))
        except Exception as e:
            logger.error(f"Error processing image {file_path}: {str(e)}")

    def process_file(self, file_path: Path) -> None:
        """
        Process a single file based on its type.
        
        Args:
            file_path (Path): Path to the file
        """
        logger.info(f"Processing {file_path}")
        
        ext = file_path.suffix.lower()
        
        if ext == '.pdf':
            self.process_pdf(file_path)
        elif ext == '.docx':
            self.process_docx(file_path)
        elif ext == '.xlsx':
            self.process_xlsx(file_path)
        elif ext in {'.jpg', '.jpeg', '.png', '.gif'}:
            self.process_image(file_path)
        elif ext == '.txt':
            # Add document title
            self.story.append(Paragraph(f"Document: {file_path.name}", self.styles['Heading1']))
            self.add_metadata_block(file_path)
            
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                if content.strip():
                    para = self.create_paragraph(content)
                    if isinstance(para, list):
                        self.story.extend(para)
                    elif para:
                        self.story.append(para)
        
        # Add a page break between documents
        self.story.append(Spacer(1, 20))

    def process_directory(self) -> None:
        """Process all supported files in the directory."""
        supported_extensions = {'.pdf', '.docx', '.xlsx', '.txt', '.jpg', '.jpeg', '.png', '.gif'}
        
        # Add title page
        self.story.append(Paragraph("Document Compilation", self.styles['Title']))
        self.story.append(Paragraph(
            f"Generated on {datetime.datetime.now().strftime('%A, %B %d, %Y at %I:%M:%S %p')}",
            self.styles['Normal']
        ))
        self.story.append(Paragraph(
            f"Source Directory: {self.input_dir}",
            self.styles['Normal']
        ))
        self.story.append(Spacer(1, 30))
        
        # Process all files
        for file_path in sorted(self.input_dir.rglob('*')):
            if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                self.process_file(file_path)

    def save_output(self) -> None:
        """Generate the final PDF."""
        doc = SimpleDocTemplate(
            str(self.output_file),
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        try:
            doc.build(self.story)
            logger.info(f"Output saved to {self.output_file}")
        except Exception as e:
            logger.error(f"Error generating PDF: {str(e)}") 